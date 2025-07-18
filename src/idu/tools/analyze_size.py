import os
import zipfile
from pathlib import Path
from collections import defaultdict
import yaml
from multiprocessing import Pool, cpu_count
from functools import partial
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# 需要忽略的文本类文件扩展名
IGNORED_EXTENSIONS = {
     '.ion'
}

# 在已经定义的 IGNORED_EXTENSIONS 下方添加需要特殊统计的非压缩包格式
SPECIAL_FORMATS = {
    '.cbz', '.cbr', '.nov'
}

def get_file_extension(filename):
    """获取文件扩展名"""
    return os.path.splitext(filename)[1].lower()  # 比 Path().suffix 更快

def analyze_zip(zip_path):
    """分析单个压缩包内的文件"""
    size_by_ext = defaultdict(int)
    file_paths = defaultdict(list)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                if not file_info.is_dir():
                    ext = get_file_extension(file_info.filename)
                    if ext not in IGNORED_EXTENSIONS:  # 过滤掉需要忽略的扩展名
                        size_by_ext[ext] += file_info.file_size
                        file_paths[ext].append(file_info.filename)
    except Exception as e:
        print(f"处理压缩包 {zip_path} 时出错: {str(e)}")
        
    return zip_path, dict(size_by_ext), dict(file_paths)

def merge_results(results, parent_dir):
    """合并多个压缩包的分析结果，按一级目录分组"""
    # 按一级目录分组的统计数据
    dir_stats = defaultdict(lambda: defaultdict(lambda: {"total_size": 0, "files": []}))
    
    for zip_path, sizes, paths in results:
        # 获取相对于父目录的一级子目录
        rel_path = os.path.relpath(zip_path, parent_dir)
        first_dir = rel_path.split(os.sep)[0]
        
        for ext, size in sizes.items():
            dir_stats[first_dir][ext]["total_size"] += size
            dir_stats[first_dir][ext]["files"].extend([
                {"zip": zip_path, "path": p} for p in paths[ext]
            ])
    
    return dict(dir_stats)

def analyze_directory(directory):
    """分析目录中的所有文件（包括非压缩包格式）"""
    size_by_ext = defaultdict(int)
    file_paths = defaultdict(list)
    
    # 遍历目录下所有文件
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            ext = get_file_extension(file)
            
            # 跳过需要忽略的文本文件格式
            if ext in IGNORED_EXTENSIONS:
                continue
                
            # 获取文件大小
            try:
                file_size = os.path.getsize(file_path)
                size_by_ext[ext] += file_size
                file_paths[ext].append(file_path)
            except (OSError, FileNotFoundError) as e:
                print(f"无法获取文件大小: {file_path}, 错误: {str(e)}")
    
    return dict(size_by_ext), dict(file_paths)

def scan_directory(directory):
    """扫描目录下的所有压缩包和特殊格式文件，按一级目录分组"""
    zip_files = []
    special_files = []
    print("正在收集文件...")
    
    # 收集压缩包文件
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_ext = get_file_extension(file)
            
            if file.lower().endswith('.zip'):
                zip_files.append(file_path)
            elif file_ext in SPECIAL_FORMATS:
                special_files.append(file_path)
    
    print(f"找到 {len(zip_files)} 个压缩包和 {len(special_files)} 个特殊格式文件")
    
    # 处理压缩包文件
    zip_results = []
    if zip_files:
        # 使用进程池并行处理压缩包
        with Pool(processes=cpu_count()) as pool:
            zip_results = list(tqdm(
                pool.imap(analyze_zip, zip_files),
                total=len(zip_files),
                desc="处理压缩包"
            ))
    
    # 处理特殊格式文件
    special_results = []
    if special_files:
        for file_path in tqdm(special_files, desc="处理特殊格式文件"):
            ext = get_file_extension(file_path)
            try:
                file_size = os.path.getsize(file_path)
                special_results.append((file_path, {ext: file_size}, {ext: [file_path]}))
            except (OSError, FileNotFoundError) as e:
                print(f"无法获取文件大小: {file_path}, 错误: {str(e)}")
    
    # 合并结果
    all_results = zip_results + special_results
    return merge_results(all_results, directory)

def format_size(size):
    """将字节大小转换为人类可读格式"""
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    size = float(size)
    unit_index = 0
    
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    
    return f"{size:.2f} {units[unit_index]}"

def calculate_total_stats(stats):
    """计算所有格式的总大小、占比和主要贡献目录"""
    total_by_ext = defaultdict(int)
    dir_sizes_by_ext = defaultdict(lambda: defaultdict(int))
    total_size = 0
    
    # 计算每种格式的总大小和各目录贡献
    for dir_name, dir_data in stats.items():
        for ext, data in dir_data.items():
            size = data["total_size"]
            total_by_ext[ext] += size
            dir_sizes_by_ext[ext][dir_name] = size
            total_size += size
    
    # 计算占比并排序
    ext_stats = {}
    for ext, size in total_by_ext.items():
        # 获取该格式下贡献最大的前5个目录
        top_dirs = sorted(
            dir_sizes_by_ext[ext].items(),
            key=lambda x: x[1],
            reverse=True
        )[:5]
        
        ext_stats[ext] = {
            "size": size,
            "percentage": (size / total_size * 100) if total_size > 0 else 0,
            "top_contributors": {
                dir_name: {
                    "size": format_size(dir_size),
                    "percentage": (dir_size / size * 100) if size > 0 else 0
                }
                for dir_name, dir_size in top_dirs
            }
        }
    
    # 按大小降序排序
    sorted_stats = dict(sorted(
        ext_stats.items(),
        key=lambda x: x[1]["size"],
        reverse=True
    ))
    
    return {
        "total_size": total_size,
        "by_extension": sorted_stats
    }

def save_results(stats, output_file):
    """保存统计结果到文件，包含总体统计和目录统计"""
    # 计算总体统计
    total_stats = calculate_total_stats(stats)
    
    # 格式化输出数据
    formatted_stats = {
        "总计": {
            "总大小": format_size(total_stats["total_size"]),
            "格式分布": {
                ext: {
                    "大小": format_size(data["size"]),
                    "占比": f"{data['percentage']:.2f}%",
                    "主要来源": {
                        dir_name: {
                            "大小": dir_data["size"],
                            "占比": f"{dir_data['percentage']:.2f}%"
                        }
                        for dir_name, dir_data in data["top_contributors"].items()
                    }
                }
                for ext, data in total_stats["by_extension"].items()
            }
        },
        "目录统计": {
            dir_name: {
                ext: format_size(data["total_size"])
                for ext, data in dir_data.items()
            }
            for dir_name, dir_data in stats.items()
        }
    }
    
    with open(output_file, 'w', encoding='utf-8') as f:
        yaml.dump(formatted_stats, f, allow_unicode=True, sort_keys=False)

def export_to_excel(stats, output_file):
    """导出统计数据到Excel并创建图表"""
    total_stats = calculate_total_stats(stats)
    wb = Workbook()
    
    # 格式统计工作表
    ws1 = wb.active
    ws1.title = "格式统计"
    
    # 写入表头
    headers = ['文件格式', '大小(GB)', '占比(%)']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 写入数据
    row = 2
    for ext, data in total_stats["by_extension"].items():
        if ext not in IGNORED_EXTENSIONS:  # 过滤掉需要忽略的扩展名
            ws1.cell(row=row, column=1, value=ext)
            ws1.cell(row=row, column=2, value=round(data["size"] / (1024**3), 1))  # 转换为GB，保留一位小数
            ws1.cell(row=row, column=3, value=round(data["percentage"], 1))
            row += 1
    
    # 创建饼图
    pie = PieChart()
    pie.title = "文件格式大小分布"
    labels = Reference(ws1, min_col=1, min_row=2, max_row=row-1)
    data = Reference(ws1, min_col=2, min_row=1, max_row=row-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws1.add_chart(pie, "E2")
    
    # 创建条形图
    bar = BarChart()
    bar.title = "前10个最大文件格式"
    bar.x_axis.title = "文件格式"
    bar.y_axis.title = "大小(GB)"
    data = Reference(ws1, min_col=2, min_row=1, max_row=min(row-1, 11))
    labels = Reference(ws1, min_col=1, min_row=2, max_row=min(row-1, 11))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    ws1.add_chart(bar, "E18")
    
    # 格式占比排名工作表
    ws2 = wb.create_sheet(title="格式占比排名")
    headers = ['文件格式', '一级目录', '大小(GB)', '目录内占比(%)', '格式总占比(%)']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 计算每个目录的总大小和每个格式的总大小
    dir_total_sizes = defaultdict(int)
    format_total_sizes = defaultdict(int)
    format_by_dir = defaultdict(lambda: defaultdict(int))
    
    for dir_name, dir_stats in stats.items():
        for ext, data in dir_stats.items():
            if ext not in IGNORED_EXTENSIONS:  # 过滤掉需要忽略的扩展名
                size = data["total_size"]
                dir_total_sizes[dir_name] += size
                format_total_sizes[ext] += size
                format_by_dir[ext][dir_name] = size
    
    # 写入格式占比排名数据
    row = 2
    current_format = None
    format_start_row = 2
    
    for ext in sorted(format_by_dir.keys()):
        # 对每个格式，按目录大小排序
        dir_sizes = sorted(
            [(dir_name, size) for dir_name, size in format_by_dir[ext].items()],
            key=lambda x: x[1],
            reverse=True
        )
        
        if current_format != ext:
            if current_format is not None:
                # 为上一个格式创建横向柱状图
                bar = BarChart()
                bar.type = "bar"  # 横向柱状图
                bar.style = 10
                bar.title = f"{current_format}格式分布"
                bar.y_axis.title = "目录"
                bar.x_axis.title = "大小(GB)"
                
                data = Reference(ws2, min_col=3, min_row=format_start_row-1, max_row=row-1)
                cats = Reference(ws2, min_col=2, min_row=format_start_row, max_row=row-1)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                
                # 计算图表位置
                chart_col = ((format_start_row - 2) // 20) * 10 + 8  # 每20行数据换一列放图表
                chart_row = ((format_start_row - 2) % 20) * 15 + 2
                ws2.add_chart(bar, f"{get_column_letter(chart_col)}{chart_row}")
            
            current_format = ext
            format_start_row = row
        
        for dir_name, size in dir_sizes:
            ws2.cell(row=row, column=1, value=ext)
            ws2.cell(row=row, column=2, value=dir_name)
            size_gb = round(size / (1024**3), 1)  # 保留一位小数
            ws2.cell(row=row, column=3, value=size_gb)
            # 计算在目录内的占比
            dir_percentage = round((size / dir_total_sizes[dir_name] * 100), 1)
            ws2.cell(row=row, column=4, value=dir_percentage)
            # 计算在该格式总大小中的占比
            format_percentage = round((size / format_total_sizes[ext] * 100), 1)
            ws2.cell(row=row, column=5, value=format_percentage)
            row += 1
    
    # 为最后一个格式创建横向柱状图
    if current_format is not None:
        bar = BarChart()
        bar.type = "bar"  # 横向柱状图
        bar.style = 10
        bar.title = f"{current_format}格式分布"
        bar.y_axis.title = "目录"
        bar.x_axis.title = "大小(GB)"
        
        data = Reference(ws2, min_col=3, min_row=format_start_row-1, max_row=row-1)
        cats = Reference(ws2, min_col=2, min_row=format_start_row, max_row=row-1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        
        chart_col = ((format_start_row - 2) // 20) * 10 + 8
        chart_row = ((format_start_row - 2) % 20) * 15 + 2
        ws2.add_chart(bar, f"{get_column_letter(chart_col)}{chart_row}")
    
    # 目录统计工作表
    ws3 = wb.create_sheet(title="目录统计")
    headers = ['目录', '总大小(GB)', '占总体比例(%)']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 计算并排序目录大小
    dir_sizes = [(dir_name, size) for dir_name, size in dir_total_sizes.items()]
    dir_sizes.sort(key=lambda x: x[1], reverse=True)
    
    row = 2
    for dir_name, size in dir_sizes:
        ws3.cell(row=row, column=1, value=dir_name)
        size_gb = round(size / (1024**3), 1)  # 保留一位小数
        ws3.cell(row=row, column=2, value=size_gb)
        percentage = round((size / total_stats["total_size"] * 100), 1)  # 保留一位小数
        ws3.cell(row=row, column=3, value=percentage)
        row += 1
    
    # 创建目录大小条形图
    bar = BarChart()
    bar.title = "目录大小分布"
    bar.x_axis.title = "目录"
    bar.y_axis.title = "大小(GB)"
    data = Reference(ws3, min_col=2, min_row=1, max_row=row-1)
    labels = Reference(ws3, min_col=1, min_row=2, max_row=row-1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    ws3.add_chart(bar, "E2")
    
    # 调整列宽
    for ws in [ws1, ws2, ws3]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存文件
    excel_path = os.path.splitext(output_file)[0] + '.xlsx'
    wb.save(excel_path)
    return excel_path

def main():
    # directory = input("请输入要扫描的目录路径：")
    # output_file = input("请输入结果保存的文件路径：")
    directory = r"E:\1EHV"
    output_file = r"E:\1EHV\size.yaml"
    
    if not os.path.exists(directory):
        print("目录不存在！")
        return
        
    print("开始扫描...")
    stats = scan_directory(directory)
    save_results(stats, output_file)
    
    print("正在生成Excel报表...")
    excel_path = export_to_excel(stats, output_file)
    print(f"扫描完成！")
    print(f"YAML结果已保存到：{output_file}")
    print(f"Excel报表已保存到：{excel_path}")
    
    # 打印简要统计信息
    print("\n按目录统计：")
    for dir_name, dir_data in stats.items():
        print(f"\n{dir_name}:")
        for ext, data in dir_data.items():
            print(f"  {ext}: {format_size(data['total_size'])}")

if __name__ == "__main__":
    main()